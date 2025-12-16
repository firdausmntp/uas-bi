"""Excel export utilities."""

from __future__ import annotations

from io import BytesIO
from typing import Mapping

import pandas as pd


def dataframes_to_xlsx_bytes(sheets: Mapping[str, pd.DataFrame]) -> bytes:
    """Build an .xlsx file (in-memory) from multiple DataFrames.

    Notes:
    - Keeps formatting lightweight but readable (header styling, freeze panes, autofilter, basic widths).
    - Requires `openpyxl`.
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = str(sheet_name)[:31] if sheet_name else "Sheet1"
            df.to_excel(writer, sheet_name=safe_name, index=False)

        # Formatting (openpyxl)
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for sheet_name, df in sheets.items():
            safe_name = str(sheet_name)[:31] if sheet_name else "Sheet1"
            ws = writer.sheets[safe_name]

            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto filter
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions

            # Header style
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Column widths (sample first 200 rows to keep it fast)
            sample_df = df.head(200) if isinstance(df, pd.DataFrame) else None
            if sample_df is not None and len(sample_df.columns) > 0:
                for col_idx, col_name in enumerate(sample_df.columns, start=1):
                    series = sample_df[col_name].astype(str)
                    max_len = max(len(str(col_name)), *(series.map(len).tolist() or [0]))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)

            # Basic number formats for sales-like columns
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_lower = str(col_name).lower()
                if "sales" in col_lower or "total" in col_lower or "avg" in col_lower:
                    col_letter = get_column_letter(col_idx)
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = "#,##0.00"

    return output.getvalue()
